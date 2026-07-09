import re
from openpyxl import load_workbook

# === CONFIGURA AQUÍ TUS RUTAS ===
INPUT_PATH = r"C:\Users\dell\Desktop\elminar\links4.xlsx"
OUTPUT_PATH = r"C:\Users\dell\Desktop\elminar\archivo_salida.xlsx"
SHEET_NAME = None  # o por ejemplo "Hoja1" si quieres una hoja específica
# =================================

CHUNK_SIZE = 60
FIGURE_PATTERN = re.compile(r'<figure\b.*?</figure>', re.DOTALL | re.IGNORECASE)


def dividir_figuras(texto):
    """Devuelve una lista de strings, cada uno con hasta CHUNK_SIZE <figure>...</figure> concatenados."""
    figuras = FIGURE_PATTERN.findall(texto)
    if not figuras:
        return [texto]
    chunks = [figuras[i:i + CHUNK_SIZE] for i in range(0, len(figuras), CHUNK_SIZE)]
    return [''.join(chunk) for chunk in chunks]


def procesar(input_path, output_path, sheet_name=None):
    wb = load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    filas_originales = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
    # Trabajamos de abajo hacia arriba para poder insertar filas sin desordenar índices
    for row in reversed(filas_originales):
        cell_a, cell_b, cell_c = row[0], row[1], row[2]
        valor_b = cell_b.value
        if not valor_b or not isinstance(valor_b, str):
            continue

        partes = dividir_figuras(valor_b)
        if len(partes) <= 1:
            continue  # 60 o menos imágenes: no se toca

        # La fila original se queda con el primer bloque (60 imágenes)
        cell_b.value = partes[0]
        titulo_base = cell_a.value
        url_c = cell_c.value
        fila_insercion = cell_a.row + 1

        for idx, parte in enumerate(partes[1:], start=2):
            ws.insert_rows(fila_insercion)
            ws.cell(row=fila_insercion, column=1, value=f"{titulo_base} #{idx}")
            ws.cell(row=fila_insercion, column=2, value=parte)
            ws.cell(row=fila_insercion, column=3, value=url_c)
            fila_insercion += 1

    wb.save(output_path)
    print(f"Listo. Archivo guardado en: {output_path}")


if __name__ == "__main__":
    procesar(INPUT_PATH, OUTPUT_PATH, SHEET_NAME)