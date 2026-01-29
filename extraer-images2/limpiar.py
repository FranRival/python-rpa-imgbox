import re
from openpyxl import load_workbook

# ================= CONFIG =================
EXCEL_FILE = "batchproceso.xlsx"

SHEETS = [
    "batch1", "batch2", "batch3", "batch4",
    "batch5", "batch6", "batch7", "batch8",
    "febrero"
]
# =========================================


def extract_urls(html_text):
    """
    Extrae todas las URLs que empiecen con:
    https://images2.imgbox.com/
    """
    if not html_text:
        return []

    pattern = r"(https://images2\.imgbox\.com/[^\s\"'>]+)"
    return re.findall(pattern, html_text)


def process_sheet(ws):
    max_col = ws.max_column
    print(f"   ➤ Columnas detectadas: {max_col}")

    for col in range(1, max_col + 1):
        html_cell = ws.cell(row=2, column=col).value
        urls = extract_urls(html_cell)

        # Borrar datos antiguos debajo (opcional pero recomendado)
        max_row = ws.max_row
        for r in range(3, max_row + 1):
            ws.cell(row=r, column=col).value = None

        if not urls:
            continue

        start_row = 3
        for i, url in enumerate(urls):
            ws.cell(row=start_row + i, column=col).value = url

    print("   ✅ Pestaña procesada.")


def main():
    print("📂 Abriendo Excel...")
    wb = load_workbook(EXCEL_FILE)

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Pestaña no encontrada: {sheet_name}")
            continue

        print(f"\n📄 Procesando pestaña: {sheet_name}")
        ws = wb[sheet_name]
        process_sheet(ws)

    print("\n💾 Guardando archivo...")
    wb.save(EXCEL_FILE)

    print("✅ Limpieza terminada correctamente.")


if __name__ == "__main__":
    main()
