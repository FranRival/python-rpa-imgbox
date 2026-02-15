from openpyxl import load_workbook, Workbook
from pathlib import Path

# -----------------------------
# CARPETA DONDE ESTÁN LOS EXCEL
# -----------------------------
SOURCE_FOLDER = Path(r"C:\Users\dell\Downloads\proceso")

# Carpeta destino (puede ser la misma si quieres)
DEST_FOLDER = Path.home() / "Desktop"

# -----------------------------
# CONSTANTES
# -----------------------------
FIGURE_TEMPLATE = (
    '<figure class="wp-block-image size-large">'
    '<img src="{url}" alt="{alt}"/>'
    '</figure>'
)

# -----------------------------
# PROCESAR TODOS LOS .xlsx
# -----------------------------
for INPUT_FILE in SOURCE_FOLDER.glob("*.xlsx"):

    print(f"Procesando: {INPUT_FILE.name}")

    wb_in = load_workbook(INPUT_FILE, data_only=True)

    # Solo Sheet1
    if "Sheet1" not in wb_in.sheetnames:
        print("  → No tiene Sheet1, se omite.")
        continue

    ws_in = wb_in["Sheet1"]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    max_col = ws_in.max_column
    max_row = ws_in.max_row

    # Copiar fila 1
    for col in range(1, max_col + 1):
        folder_name = ws_in.cell(row=1, column=col).value
        ws_out.cell(row=1, column=col).value = folder_name

    # Construir HTML en fila 2
    for col in range(1, max_col + 1):
        folder_name = ws_in.cell(row=1, column=col).value

        if not folder_name:
            continue

        html_parts = []

        for row in range(3, max_row + 1):
            url = ws_in.cell(row=row, column=col).value
            if not url:
                continue

            html_parts.append(
                FIGURE_TEMPLATE.format(
                    url=str(url).strip(),
                    alt=str(folder_name).strip()
                )
            )

        ws_out.cell(row=2, column=col).value = "".join(html_parts)

    # Guardar con mismo nombre
    OUTPUT_FILE = DEST_FOLDER / INPUT_FILE.name
    wb_out.save(OUTPUT_FILE)

    print(f"  → Guardado como: {OUTPUT_FILE.name}")

print("Proceso completado.")
