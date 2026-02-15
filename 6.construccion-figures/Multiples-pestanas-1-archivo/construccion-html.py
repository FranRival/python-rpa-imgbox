from openpyxl import load_workbook, Workbook
from pathlib import Path

# -----------------------------
# RUTAS
# -----------------------------
INPUT_FILE = Path(r"C:\Users\dell\Downloads\proceso\10.xlsx")
OUTPUT_FILE = Path.home() / "Desktop" / "estructura.xlsx"

# -----------------------------
# CONSTANTES
# -----------------------------
IGNORE_SHEETS = {"basura"}

FIGURE_TEMPLATE = (
    '<figure class="wp-block-image size-large">'
    '<img src="{url}" alt="{alt}"/>'
    '</figure>'
)

# -----------------------------
# CARGA EXCEL ORIGEN
# -----------------------------
wb_in = load_workbook(INPUT_FILE, data_only=True)

# -----------------------------
# CREAR EXCEL DESTINO
# -----------------------------
wb_out = Workbook()
# eliminar hoja por defecto
default_sheet = wb_out.active
wb_out.remove(default_sheet)

# -----------------------------
# PROCESAR CADA PESTAÑA
# -----------------------------
for sheet_name in wb_in.sheetnames:
    if sheet_name.lower() in IGNORE_SHEETS:
        continue

    ws_in = wb_in[sheet_name]
    ws_out = wb_out.create_sheet(title=sheet_name)

    max_col = ws_in.max_column
    max_row = ws_in.max_row

    # -------------------------
    # FILA 1: copiar nombres de carpetas TAL CUAL
    # -------------------------
    for col in range(1, max_col + 1):
        folder_name = ws_in.cell(row=1, column=col).value
        ws_out.cell(row=1, column=col).value = folder_name

    # -------------------------
    # FILA 2: construir HTML
    # -------------------------
    for col in range(1, max_col + 1):
        folder_name = ws_in.cell(row=1, column=col).value

        if not folder_name:
            continue

        html_parts = []

        # recorrer URLs desde fila 3 hacia abajo
        for row in range(3, max_row + 1):
            url = ws_in.cell(row=row, column=col).value
            if not url:
                continue

            html_parts.append(
                FIGURE_TEMPLATE.format(
                    url=str(url).strip(),
                    alt=str(folder_name).strip()
                )
            )

        # todo el HTML en UNA sola celda, sin saltos de línea
        ws_out.cell(row=2, column=col).value = "".join(html_parts)

# -----------------------------
# GUARDAR RESULTADO
# -----------------------------
wb_out.save(OUTPUT_FILE)

print("Proceso completado.")
print(f"Archivo generado en: {OUTPUT_FILE}")
