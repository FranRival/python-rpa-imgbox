import re
import os
from openpyxl import load_workbook

# ================= CONFIG =================
FOLDER_PATH = r"C:\Users\dell\Desktop\uploader\5.extraer-images2"   # Carpeta actual (puedes poner ruta absoluta si quieres)
TARGET_SHEET = "Sheet1"
# ==========================================


def extract_urls(html_text):
    """
    Extrae todas las URLs que empiecen con:
    https://images2.imgbox.com/
    """
    if not html_text:
        return []

    pattern = r"(https://images2\.imgbox\.com/[^\s\"'>]+)"
    return re.findall(pattern, html_text)


def process_sheet(ws):
    max_col = ws.max_column
    print(f"   ➤ Columnas detectadas: {max_col}")

    for col in range(1, max_col + 1):
        html_cell = ws.cell(row=2, column=col).value
        urls = extract_urls(html_cell)

        # Borrar datos antiguos debajo
        max_row = ws.max_row
        for r in range(3, max_row + 1):
            ws.cell(row=r, column=col).value = None

        if not urls:
            continue

        start_row = 3
        for i, url in enumerate(urls):
            ws.cell(row=start_row + i, column=col).value = url

    print("   ✅ Pestaña procesada.")


def process_file(file_path):
    print(f"\n📂 Procesando archivo: {file_path}")
    wb = load_workbook(file_path)

    if TARGET_SHEET not in wb.sheetnames:
        print(f"⚠️ No se encontró la pestaña {TARGET_SHEET}")
        return

    ws = wb[TARGET_SHEET]
    process_sheet(ws)

    wb.save(file_path)
    print("💾 Archivo guardado.")


def main():
    print("🔍 Buscando archivos Excel...")

    for filename in os.listdir(FOLDER_PATH):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(FOLDER_PATH, filename)
            process_file(file_path)

    print("\n✅ Todos los archivos procesados.")


if __name__ == "__main__":
    main()
