import re
from openpyxl import load_workbook
import sys
import os

ARCHIVO_EXCEL = "imagenes.xlsx"  # nombre exacto del archivo en la misma carpeta

def extraer_urls(html):
    if not html:
        return []
    return re.findall(r'https://images\d\.imgbox\.com/[^\s"\']+\.(?:jpg|png|gif)', html)

def generar_figure(urls, titulo):
    bloques = []
    for url in urls:
        bloques.append(f'<figure class="wp-block-image size-large"><img src="{url}" alt="{titulo}"/></figure>')
    return "".join(bloques)

# ==== Verificar que el archivo existe en la carpeta actual ====
if not os.path.exists(ARCHIVO_EXCEL):
    print(f"ERROR: no se encontró '{ARCHIVO_EXCEL}' en la carpeta actual: {os.getcwd()}")
    sys.exit(1)

# ==== Abrir el workbook y mostrar hojas disponibles ====
wb = load_workbook(ARCHIVO_EXCEL)
sheet_names = wb.sheetnames
print("Hojas detectadas en el Excel:", sheet_names)

# ==== Seleccionar hoja: usar la primera hoja por defecto ====
ws = wb[sheet_names[0]]
print(f"Usando la hoja: '{ws.title}'\n")

# ==== Procesar todas las columnas con datos en fila 2 ====
procesadas = 0
for col in ws.iter_cols(min_row=2, max_row=2):
    celda_html = col[0]  # celda en fila 2
    col_letra = celda_html.column_letter
    html = celda_html.value

    if not html:
        # saltar si la celda está vacía
        print(f"- Columna {col_letra}: vacía, saltando.")
        continue

    nombre_titulo = ws[f"{col_letra}1"].value or "IMAGEN"
    urls = extraer_urls(html)
    codigo_limpio = generar_figure(urls, nombre_titulo)

    ws[f"{col_letra}3"] = codigo_limpio
    print(f"- Columna {col_letra}: {len(urls)} imágenes procesadas.")
    procesadas += 1

# ==== Guardar workbook ====
wb.save(ARCHIVO_EXCEL)
print(f"\n✔ Terminado. {procesadas} columnas procesadas. El resultado está en la fila 3 de cada columna.")
