from openpyxl import load_workbook

wb = load_workbook("imagenes.xlsx")
print(wb.sheetnames)
