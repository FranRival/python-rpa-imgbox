import os
import sys
import time
import traceback
from pathlib import Path

# 🔥 EVITA BLOQUEOS DE CMD / .BAT
sys.stdin = open(os.devnull)

# 🔥 EVITA QUE CMD "TRAGUE" LOS PRINTS (buffering) — así se ven en tiempo real
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except Exception:
    pass

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook, load_workbook


# =========================
# CONFIGURACIÓN
# =========================

FREEIMAGE_URL = "https://freeimage.host/es-mx"
ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}  # freeimage.host no acepta .mp4

# Dominios válidos que puede devolver el HTML generado (freeimage.host usa iili.io como CDN)
DOMINIOS_VALIDOS = ("freeimage.host", "iili.io")


# =========================
# UTILIDADES
# =========================

def obtener_subcarpetas(root):
    return [
        os.path.join(root, d)
        for d in sorted(os.listdir(root))
        if os.path.isdir(os.path.join(root, d))
    ]


# =========================
# EXCEL (1 ARCHIVO POR BATCH)
# =========================

def get_excel_path(batch_root):
    batch_name = os.path.basename(batch_root.rstrip("\\/"))
    return os.path.join(batch_root, f"{batch_name}.xlsx")


def ensure_excel(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(path)
    return load_workbook(path)


def find_or_create_column(ws, header):
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        if (ws.cell(row=1, column=col).value or "").strip() == header:
            return col

    new_col = max_col + 1
    ws.cell(row=1, column=new_col, value=header)
    return new_col


def write_html_to_excel(excel_path, folder_name, html):
    wb = ensure_excel(excel_path)
    ws = wb["Sheet1"]

    col = find_or_create_column(ws, folder_name)
    ws.cell(row=2, column=col, value=html)

    wb.save(excel_path)
    print(f"💾 Excel actualizado → {excel_path} | Columna: {folder_name}")


# =========================
# SELENIUM
# =========================

def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    # Evita el diálogo nativo "guardar contraseña" / notificaciones que puedan tapar botones
    options.add_experimental_option("prefs", {
        "profile.default_content_setting_values.notifications": 2
    })
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(60)
    return driver


def cerrar_popups(driver):
    """Cierra banners de cookies u overlays que puedan bloquear clics."""
    posibles_textos = ["Aceptar", "Accept", "OK", "Entendido"]
    for texto in posibles_textos:
        try:
            btn = driver.find_element(
                By.XPATH, f"//button[contains(normalize-space(.),'{texto}')]"
            )
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(1)
        except NoSuchElementException:
            pass


# =========================
# IDs REALES confirmados a partir del HTML fuente de freeimage.host
# (motor Chevereto). Nada de esto se adivina: son los ids/valores exactos.
# =========================
ID_INPUT_ARCHIVOS = "anywhere-upload-input"          # <input type="file"> real
SELECTOR_BOTON_SUBIR = "[data-action='upload']"       # <button data-action="upload">Subir</button>
SELECTOR_RESULTADO_EXITO = "[data-group='upload-result'][data-result='success']"
SELECTOR_RESULTADO_ERROR = "[data-group='upload-result'][data-result='error']"
ID_SELECT_CODIGOS = "uploaded-embed-toggle"           # <select> real del resultado (NO el del modal "form-embed-toggle")
VALUE_HTML_MEDIO = "html-embed-medium"                # value de "HTML completo con enlace"
ID_TEXTAREA_HTML_MEDIO = "uploaded-embed-code-1"      # <textarea> que siempre contiene ese formato


# =========================
# PASO 1: SELECCIONAR ARCHIVOS
# =========================

def seleccionar_archivos(driver, archivos):
    wait = WebDriverWait(driver, 30)
    input_file = wait.until(
        EC.presence_of_element_located((By.ID, ID_INPUT_ARCHIVOS))
    )
    # No hace falta que sea visible para send_keys en Selenium con Chrome
    input_file.send_keys("\n".join(archivos))
    time.sleep(2)


# =========================
# PASO 2: CLIC EN BOTÓN VERDE "SUBIR"
# =========================

def click_boton_subir(driver):
    wait = WebDriverWait(driver, 30)
    # <button data-action="upload" ...>Subir</button> — atributo único y estable,
    # no depende del idioma ni del texto visible.
    boton = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, SELECTOR_BOTON_SUBIR))
    )
    driver.execute_script("arguments[0].click();", boton)


# =========================
# PASO 3: ESPERAR "SUBIDA COMPLETA"
# =========================

def esperar_subida_completa(driver, timeout=300):
    """
    El sitio controla el estado mostrando/ocultando divs mediante la clase
    'soft-hidden' (que en su CSS es literalmente display:none). El resultado
    exitoso vive en:
        <div data-group="upload-result" data-result="success" class="soft-hidden">
    Cuando la subida termina, el sitio le QUITA la clase 'soft-hidden' a este
    div. Chequeamos eso directamente (is_displayed) en vez de buscar texto.
    """
    print("⌛ Esperando a que la subida termine (Subida completa)...")
    fin = time.time() + timeout
    intentos = 0
    while time.time() < fin:
        exitos = driver.find_elements(By.CSS_SELECTOR, SELECTOR_RESULTADO_EXITO)
        if any(e.is_displayed() for e in exitos):
            print("✅ Subida completa detectada")
            return

        errores = driver.find_elements(By.CSS_SELECTOR, SELECTOR_RESULTADO_ERROR)
        visibles_error = [e for e in errores if e.is_displayed()]
        if visibles_error:
            texto_error = visibles_error[0].text.strip()
            raise RuntimeError(f"El sitio reportó un error de subida: {texto_error!r}")

        intentos += 1
        if intentos % 15 == 0:
            print(f"   ...sigue subiendo ({intentos * 2}s transcurridos)")
        time.sleep(2)

    raise TimeoutException("La subida no terminó (no se detectó 'Subida completa' visible) a tiempo.")


# =========================
# PASO 4: ELEGIR "HTML completo con enlace" EN EL SELECT
# =========================

def seleccionar_html_completo(driver, timeout=30):
    """
    <select id="uploaded-embed-toggle"> es el select REAL del panel de
    resultados (distinto del <select id="form-embed-toggle"> que pertenece a
    un modal aparte que no se usa en este flujo). Su value para el formato
    que queremos es 'html-embed-medium'.
    """
    print(f"🔍 Buscando el selector '{ID_SELECT_CODIGOS}'...")
    fin = time.time() + timeout
    select_el = None

    while time.time() < fin:
        candidatos = driver.find_elements(By.ID, ID_SELECT_CODIGOS)
        if candidatos:
            select_el = candidatos[0]
            break
        time.sleep(0.5)

    if select_el is None:
        raise TimeoutException(f"No se encontró el <select id='{ID_SELECT_CODIGOS}'> a tiempo.")

    print(f"✅ Select encontrado. Fijando value='{VALUE_HTML_MEDIO}'...")
    driver.execute_script(
        """
        var select = arguments[0];
        select.value = arguments[1];
        select.dispatchEvent(new Event('change', { bubbles: true }));
        """,
        select_el,
        VALUE_HTML_MEDIO,
    )
    time.sleep(1)


# =========================
# PASO 5: EXTRAER EL HTML GENERADO
# =========================

def extraer_html_generado(driver, timeout=60):
    """
    El textarea con id='uploaded-embed-code-1' SIEMPRE contiene el código
    'HTML completo con enlace' (así el <select> no esté en esa opción, el
    combo ya viene pre-armado en el DOM para cada formato — confirmado en el
    HTML fuente). Por eso lo leemos directo por id, sin adivinar patrones.
    """
    print("⌛ Esperando el HTML en el textarea...")
    fin = time.time() + timeout
    while time.time() < fin:
        try:
            area = driver.find_element(By.ID, ID_TEXTAREA_HTML_MEDIO)
            val = (area.get_attribute("value") or "").strip()
            if val:
                print("✅ HTML detectado")
                return val
        except NoSuchElementException:
            pass
        time.sleep(1)
    raise TimeoutException(f"No se detectó contenido en el <textarea id='{ID_TEXTAREA_HTML_MEDIO}'> a tiempo.")


# =========================
# PROCESO DE SUBIDA POR CARPETA
# =========================

def subir_carpeta(driver, excel_path, folder):
    archivos = [
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if Path(f).suffix.lower() in ALLOWED_EXT
    ]

    if not archivos:
        print(f"⚠️ Carpeta vacía: {folder}")
        return

    nombre_carpeta = os.path.basename(folder)

    print("\n📁 SUBCARPETA ACTUAL (RUTA COMPLETA):")
    print(folder)
    print(f"📦 Archivos: {len(archivos)}")

    driver.get(FREEIMAGE_URL)
    time.sleep(3)
    cerrar_popups(driver)

    seleccionar_archivos(driver, archivos)
    click_boton_subir(driver)
    esperar_subida_completa(driver)
    seleccionar_html_completo(driver)

    html = extraer_html_generado(driver)
    write_html_to_excel(excel_path, nombre_carpeta, html)


# =========================
# MAIN
# =========================

def main():
    if len(sys.argv) < 2:
        print("❌ No se recibió ruta del batch")
        sys.exit(1)

    batch_root = sys.argv[1]

    print("\n🚀 INICIANDO UPLOADER (freeimage.host)")
    print("📂 RUTA COMPLETA DEL BATCH:")
    print(batch_root)

    excel_path = get_excel_path(batch_root)
    print(f"📊 Excel del batch: {excel_path}")

    carpetas = obtener_subcarpetas(batch_root)
    print(f"📁 Total carpetas: {len(carpetas)}")

    driver = init_driver()

    try:
        for idx, carpeta in enumerate(carpetas, 1):
            print(f"\n➡️ {idx}/{len(carpetas)}")
            try:
                subir_carpeta(driver, excel_path, carpeta)
            except Exception:
                print(f"\n❌ ERROR en la carpeta: {carpeta}")
                traceback.print_exc()
                continue  # sigue con la siguiente carpeta en vez de morir todo el batch

        print("\n🏁 BATCH COMPLETADO")

    except Exception:
        print("\n❌ ERROR CRÍTICO")
        traceback.print_exc()

    finally:
        time.sleep(5)
        driver.quit()
        print("✅ Proceso terminado")


if __name__ == "__main__":
    main()